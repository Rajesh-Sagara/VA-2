"""
Excel / CSV Parser – reads spreadsheet content and produces per-sheet hashes.
Each sheet's data is serialised to CSV, then SHA-256 hashed independently.
"""

import io
import hashlib
import csv


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def parse_excel(content: bytes, filename: str = "") -> dict:
    """
    Parse an Excel (.xlsx/.xls) or CSV file.

    Returns:
        {
            "format": "xlsx" | "xls" | "csv",
            "sheets": [
                {
                    "name": str,
                    "row_count": int,
                    "col_count": int,
                    "hash": str,          # SHA-256 of the sheet's CSV text
                    "preview": [[...]]    # first 5 rows
                }
            ],
            "total_sheets": int,
            "error": str | None
        }
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # ── CSV ──────────────────────────────────────────────────────────────────
    if ext == "csv":
        try:
            text = content.decode("utf-8", errors="replace")
            reader = list(csv.reader(io.StringIO(text)))
            rows = len(reader)
            cols = max((len(r) for r in reader), default=0)
            sheet_hash = _hash_bytes(text.encode())
            return {
                "format": "csv",
                "sheets": [{
                    "name": "Sheet1",
                    "row_count": rows,
                    "col_count": cols,
                    "hash": sheet_hash,
                    "preview": reader[:5],
                }],
                "total_sheets": 1,
                "error": None,
            }
        except Exception as e:
            return {"format": "csv", "sheets": [], "total_sheets": 0, "error": str(e)}

    # ── XLSX / XLS ────────────────────────────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                rows_data.append([str(c) if c is not None else "" for c in row])

            row_count = len(rows_data)
            col_count = max((len(r) for r in rows_data), default=0)

            # Serialise to CSV bytes for hashing
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerows(rows_data)
            sheet_bytes = buf.getvalue().encode()
            sheet_hash = _hash_bytes(sheet_bytes)

            sheets.append({
                "name": sheet_name,
                "row_count": row_count,
                "col_count": col_count,
                "hash": sheet_hash,
                "preview": rows_data[:5],
            })

        return {
            "format": "xlsx",
            "sheets": sheets,
            "total_sheets": len(sheets),
            "error": None,
        }
    except ImportError:
        return {
            "format": ext,
            "sheets": [],
            "total_sheets": 0,
            "error": "openpyxl not installed. Run: pip install openpyxl",
        }
    except Exception as e:
        return {"format": ext, "sheets": [], "total_sheets": 0, "error": str(e)}
